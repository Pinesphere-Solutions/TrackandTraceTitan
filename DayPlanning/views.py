from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from modelmasterapp.models import *  # Import the TrayId model
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from modelmasterapp.models import *
from modelmasterapp.models import *
from django.contrib import messages
from datetime import datetime
from django.utils.safestring import mark_safe
import re
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.parsers import MultiPartParser
from rest_framework import status
import openpyxl
import re
from datetime import datetime
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.db.models import OuterRef, Subquery
import json
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.parsers import MultiPartParser, JSONParser
import math
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import TemplateHTMLRenderer
import math
import json
from rest_framework.permissions import IsAuthenticated


class DPBulkUploadView(APIView):
    """
    Enhanced Day Planning Bulk Upload View
    
    Handles two source formats:
    1. Vendor_Location format (e.g., "Titan_CPSE") - splits into vendor and location
    2. Location only format (e.g., "CPSE") - treats entire value as location name
    
    Supports ambiguous plating color resolution using additional identifiers
    after "/" in plating stock numbers (e.g., "2648QAA02/BRN").
    
    Enhanced with comprehensive column validation and detailed error reporting.
    """
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_BulkUpload.html'
    parser_classes = [MultiPartParser, JSONParser]
    permission_classes = [IsAuthenticated] 

    def validate_excel_columns(self, sheet):
        """
        Validate Excel file columns to ensure they match expected format
        Returns: (is_valid, error_message, header_row)
        """
        try:
            # Expected column names in exact order
            expected_columns = [
                'S.No',
                'Plating Stk No', 
                'Polishing Stk No',
                'Plating Colour',
                'Category',
                'Input Qty',
                'Source'
            ]
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_count = len(expected_columns) - len(actual_columns)
                missing_columns = expected_columns[len(actual_columns):]
                return False, f"❌ Missing {missing_count} column(s): {', '.join(missing_columns)}. Expected {len(expected_columns)} columns but found {len(actual_columns)}.", actual_columns
            
            # Check for exact column name matches
            mismatched_columns = []
            for i, (expected, actual) in enumerate(zip(expected_columns, actual_columns)):
                if expected != actual:
                    mismatched_columns.append(f"Column {i+1}: Expected '{expected}' but found '{actual}'")
            
            if mismatched_columns:
                error_msg = "❌ Column Name Mismatch:\n" + "\n".join(mismatched_columns)
                error_msg += f"\n\nExpected format: {' | '.join(expected_columns)}"
                error_msg += f"\nActual format: {' | '.join(actual_columns[:len(expected_columns)])}"
                return False, error_msg, actual_columns
            
            # Check for extra columns (warn but don't fail)
            if len(actual_columns) > len(expected_columns):
                extra_columns = actual_columns[len(expected_columns):]
                print(f"⚠️ Warning: Found {len(extra_columns)} extra column(s) that will be ignored: {', '.join(extra_columns)}")
                
            
            return True, "✅ Column validation passed", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None




    def validate_codes(self, plating_stock_no, polishing_stock_no):
        """
        Enhanced validation to handle ambiguous plating codes.
        
        When a plating code (like 'Q') maps to multiple colors in the database,
        this method looks for additional identifier after "/" in plating_stock_no.
        
        Example: "2648QAA02/BRN" 
        - Q is the plating code but ambiguous
        - /BRN is the specific identifier to match plating_color_internal='BRN'
        
        IMPORTANT: Stock number validation rules:
        - Plating: 1805SAB02 (any UPPERCASE color code, must end with 02)
        - Polishing: 1805XAB02 (must have UPPERCASE X as color code, must end with 02)
        - Model numbers must match: 1805 = 1805
        - ALL LETTERS MUST BE UPPERCASE: SAB02 ✅, sab02 ❌, SAb02 ❌
        - Suffixes must match except color: SAB02 vs XAB02 (only S vs X difference)
        
        Returns: (model_stock, codes_tuple, error_msg)
        """
        try:
            # STEP 0: Validate that both stock numbers end with "02"
            plating_base = plating_stock_no.split("/")[0] if "/" in plating_stock_no else plating_stock_no
            polishing_base = polishing_stock_no.split("/")[0] if "/" in polishing_stock_no else polishing_stock_no
            
            if not plating_base.endswith("02"):
                return None, None, f"❌ Invalid Plating Stk No format: '{plating_stock_no}' must end with '02'. Example: 1805SAB02"
            
            if not polishing_base.endswith("02"):
                return None, None, f"❌ Invalid Polishing Stk No format: '{polishing_stock_no}' must end with '02'. Example: 1805XAB02"

            # STEP 0.1: CRITICAL - Validate that ALL letters are UPPERCASE
            plating_letters = ''.join(re.findall(r'[A-Za-z]', plating_base))
            if plating_letters != plating_letters.upper():
                lowercase_found = [c for c in plating_letters if c.islower()]
                return None, None, f"❌ Invalid Plating Stk No: '{plating_stock_no}' contains lowercase letters '{', '.join(lowercase_found)}'. ALL letters must be UPPERCASE. Correct format: {plating_base.upper()}"
            
            polishing_letters = ''.join(re.findall(r'[A-Za-z]', polishing_base))
            if polishing_letters != polishing_letters.upper():
                lowercase_found = [c for c in polishing_letters if c.islower()]
                return None, None, f"❌ Invalid Polishing Stk No: '{polishing_stock_no}' contains lowercase letters '{', '.join(lowercase_found)}'. ALL letters must be UPPERCASE. Correct format: {polishing_base.upper()}"

            # STEP 1: Validate that plating and polishing stock numbers match pattern
            # Extract model number and pattern from both stock numbers (ONLY UPPERCASE)
            plating_match = re.match(r'^(\d+)([A-Z])([A-Z][A-Z]02)$', plating_base)
            polishing_match = re.match(r'^(\d+)([A-Z])([A-Z][A-Z]02)$', polishing_base)
            
            if not plating_match:
                return None, None, f"❌ Invalid Plating Stk No format: '{plating_stock_no}'. Expected format: ModelNumber + UPPERCASE_ColorCode + UPPERCASE_Letters + 02 (e.g., 1805SAB02). Current: {plating_base}"
            
            if not polishing_match:
                return None, None, f"❌ Invalid Polishing Stk No format: '{polishing_stock_no}'. Expected format: ModelNumber + X + UPPERCASE_Letters + 02 (e.g., 1805XAB02). Current: {polishing_base}"
            
            plating_model, plating_color_code, plating_suffix = plating_match.groups()
            polishing_model, polishing_color_code, polishing_suffix = polishing_match.groups()
            
            # STEP 1.1: Validate that polishing stock number has "X" as color code (UPPERCASE)
            if polishing_color_code != 'X':
                return None, None, f"❌ Invalid Polishing Stk No: '{polishing_stock_no}'. Polishing stock number must have UPPERCASE 'X' as color code. Expected: {polishing_model}X{polishing_suffix} (not '{polishing_color_code}')"
            
            # STEP 1.2: Validate that model numbers match
            if plating_model != polishing_model:
                return None, None, f"❌ Stock number mismatch: Plating model '{plating_model}' does not match Polishing model '{polishing_model}'. Expected format: {plating_model}[COLOR]AB02 for plating and {plating_model}XAB02 for polishing."
            
            # STEP 1.3: Create expected polishing suffix by replacing color code with X
            expected_polishing_suffix = plating_suffix
            expected_polishing_stock = f"{plating_model}X{expected_polishing_suffix}"
            
            # Validate that polishing suffix matches plating suffix (except for color code)
            if polishing_suffix != expected_polishing_suffix:
                return None, None, f"❌ Stock number mismatch: Expected polishing stock number '{expected_polishing_stock}' but got '{polishing_stock_no}'. Only the color code should be 'X'."
            
            # STEP 1.4: Additional validation: Ensure suffix follows UPPERCASE AB02 pattern
            if not re.match(r'^[A-Z][A-Z]02$', plating_suffix):
                return None, None, f"❌ Invalid suffix pattern in Plating Stk No: '{plating_suffix}'. Expected pattern: [UPPERCASE][UPPERCASE]02 (e.g., AB02, not ab02 or Ab02)"
            
            if not re.match(r'^[A-Z][A-Z]02$', polishing_suffix):
                return None, None, f"❌ Invalid suffix pattern in Polishing Stk No: '{polishing_suffix}'. Expected pattern: [UPPERCASE][UPPERCASE]02 (e.g., AB02, not ab02 or Ab02)"
            
            # The polishing stock should have X and match the plating pattern
            print(f"✅ Stock number validation passed:")
            print(f"   Plating: {plating_model}{plating_color_code}{plating_suffix}")
            print(f"   Polishing: {polishing_model}X{polishing_suffix} ✓")

            # STEP 2: Extract model number and validate in ModelMaster
            model_no = plating_model
            model_stock = ModelMaster.objects.filter(model_no=model_no).first()
            if not model_stock:
                return None, None, f"❌ Plating Stk No '{plating_stock_no}' - Model number '{model_no}' not available in Master Data."

            print(f"✅ Model number '{model_no}' exists in ModelMaster.")

            # STEP 3: Determine plating color internal code
            plating_color_internal = None
            if "/" in plating_stock_no:
                # Extract the part after "/" for ambiguous resolution
                additional_identifier = plating_stock_no.split("/")[1]
                plating_color_internal = additional_identifier
                print(f"🔍 Found additional identifier after '/': {additional_identifier} (resolving ambiguous plating code)")
            else:
                # Use the plating color code directly
                plating_color_internal = plating_color_code
                print(f"🔍 Using plating color code directly: {plating_color_code}")

            print(f"📋 Final plating_color_internal to lookup: '{plating_color_internal}'")

            # STEP 4: Extract polish and version codes from polishing stock number
            # For polishing, we know the color code is X, so we extract from the known pattern
            letters = ''.join(re.findall(r'[A-Z]', polishing_stock_no))  # Only UPPERCASE letters
            if len(letters) < 3:
                return None, None, f"❌ Invalid polishing stock format. Found less than 3 UPPERCASE letters in: {polishing_stock_no}. Expected format: ModelNumber + X + [UPPERCASE][UPPERCASE]02"

            # Since we know polishing has X as color code, extract polish and version codes
            polish_code = letters[1]  # Second letter for polish (after X)
            version_code = letters[2]  # Third letter for version

            # Return as tuple of 3 values: model_stock, codes_tuple, error_msg
            codes_tuple = (plating_color_internal, polish_code, version_code)
            return model_stock, codes_tuple, None

        except Exception as e:
            return None, None, f"❌ Error validating codes: {str(e)}"

    def get(self, request, format=None):
        master_data = ModelMasterCreation.objects.all()
        return Response({'master_data': master_data})

    def post(self, request, format=None):
        # Check if this is a preview request
        if 'preview' in request.path:
            return self.handle_file_preview(request)
        
        # Check if request contains JSON data (from datatable)
        if request.content_type == 'application/json':
            return self.handle_datatable_submission(request)
        
        # Handle file upload (existing functionality)
        return self.handle_file_upload(request)

    def handle_file_preview(self, request):
        """Handle file preview for datatable with column validation"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'error': '❌ No file uploaded.'
            }, status=400)

        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return JsonResponse({
                'success': False,
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            if not sheet:
                return JsonResponse({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # ========== VALIDATE COLUMNS FIRST ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                return JsonResponse({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            print("✅ Excel column validation passed for preview")
            # ========== END COLUMN VALIDATION ==========

            # Process rows for preview
            preview_data = []
            row_errors = []
            
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                # Check for minimum required columns
                if len(row) < 7:
                    row_errors.append(f"Row {idx}: ❌ Missing Data — Expected 7 columns but found {len(row)}")
                    continue

                # Extract data
                s_no = str(row[0]).strip() if row[0] else ''
                plating_stock_no = str(row[1]).strip() if row[1] else ''
                polishing_stock_no = str(row[2]).strip() if row[2] else ''
                plating_colour = str(row[3]).strip() if row[3] else ''
                category = str(row[4]).strip() if row[4] else ''
                input_qty = row[5]
                source = str(row[6]).strip() if row[6] else ''

                # Basic validation for preview
                row_data = {
                    'S.No': s_no or str(len(preview_data) + 1),
                    'Plating Stk No': plating_stock_no,
                    'Polishing Stk No': polishing_stock_no,
                    'Plating Colour': plating_colour,
                    'Category': category,
                    'Input Qty': input_qty,
                    'Source': source
                }
                
                preview_data.append(row_data)
                
                # Limit preview to first 100 rows
                if len(preview_data) >= 100:
                    break

            if not preview_data and row_errors:
                return JsonResponse({
                    'success': False,
                    'error': f"❌ No valid data found.\n\n" + "\n".join(row_errors[:5])
                }, status=400)

            response_data = {
                'success': True,
                'data': preview_data
            }
            
            if row_errors:
                response_data['warnings'] = row_errors[:10]  # Show first 10 warnings
                
            return JsonResponse(response_data)

        except Exception as e:
            print(f"❌ Error processing file preview: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'❌ An error occurred: {str(e)}'
            }, status=500)

    def handle_datatable_submission(self, request):
        """Handle submission from HTML datatable"""
        try:
            # Parse JSON data from request body
            data = request.data
            rows = data.get('rows', [])
            
            if not rows:
                return JsonResponse({
                    'success': False,
                    'error': '❌ No data provided for processing.'
                }, status=400)

            success_count = 0
            failure_count = 0
            failed_rows = []

            for idx, row_data in enumerate(rows, start=1):
                try:
                    # Extract data from row with new field names
                    s_no = str(row_data.get('S.No', '')).strip()
                    plating_stock_no = str(row_data.get('Plating Stk No', '')).strip()
                    polishing_stock_no = str(row_data.get('Polishing Stk No', '')).strip()
                    plating_colour = str(row_data.get('Plating Colour', '')).strip()
                    category = str(row_data.get('Category', '')).strip()
                    input_qty = row_data.get('Input Qty')
                    source = str(row_data.get('Source', '')).strip()

                    print(f"\n🔄 Processing row {idx}: Plating Stk No={plating_stock_no}, Polishing Stk No={polishing_stock_no}, Colour={plating_colour}, Category={category}, Qty={input_qty}, Source={source}")

                    # Validate required fields
                    empty_fields = []
                    if not plating_stock_no:
                        empty_fields.append("Plating Stk No")
                    if not polishing_stock_no:
                        empty_fields.append("Polishing Stk No")
                    if not plating_colour:
                        empty_fields.append("Plating Colour")
                    if not category:
                        empty_fields.append("Category")
                    if input_qty in [None, '', 0]:
                        empty_fields.append("Input Qty")
                    if not source:
                        empty_fields.append("Source")

                    if empty_fields:
                        field_list = ", ".join(empty_fields)
                        failed_rows.append(f"Row {idx}: ❌ {field_list} should not be empty.")
                        failure_count += 1
                        continue

                    # Convert quantity to integer
                    try:
                        input_qty = int(input_qty)
                    except (ValueError, TypeError):
                        failed_rows.append(f"Row {idx}: ❌ Invalid quantity value: {input_qty}")
                        failure_count += 1
                        continue

                    # Validate codes using new stock numbers - FIXED UNPACKING
                    model_stock, codes, error_msg = self.validate_codes(plating_stock_no, polishing_stock_no)
                    if error_msg:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: {error_msg}")
                        continue

                    if not codes:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Could not extract codes from stock numbers.")
                        continue

                    plating_color_internal, polish_code, version_code = codes
                    print(f"🧪 Codes Extracted ➤ Plating Internal: '{plating_color_internal}', Polish: '{polish_code}', Version: '{version_code}'")

                    # ENHANCED VALIDATION WITH DETAILED ERROR MESSAGES

                    # 1. Validate plating code using the resolved plating_color_internal
                    plating_obj_code = Plating_Color.objects.filter(plating_color_internal=plating_color_internal).first()
                    if not plating_obj_code:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating color internal code '{plating_color_internal}' not available in Master Data.")
                        continue

                    # 2. Validate plating color from input
                    plating_color_obj = Plating_Color.objects.filter(plating_color=plating_colour).first()
                    if not plating_color_obj:
                        failure_count += 1
                        # Get available plating colors for suggestion
                        available_colors = list(Plating_Color.objects.values_list('plating_color', flat=True)[:5])
                        color_suggestion = f" Available colors: {', '.join(available_colors)}" if available_colors else ""
                        failed_rows.append(f"Row {idx}: ❌ Plating Colour '{plating_colour}' not available in Master Data.{color_suggestion}")
                        continue

                    # 3. Cross-validation: Check if resolved plating_color_internal matches with the plating color from Excel
                    if plating_obj_code.pk != plating_color_obj.pk:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating color mismatch: Stock code '{plating_stock_no}' resolves to '{plating_obj_code.plating_color}' but Excel shows '{plating_colour}'.")
                        continue

                    # 4. Validate polish code
                    polish_obj = PolishFinishType.objects.filter(polish_internal=polish_code).first()
                    if not polish_obj:
                        failure_count += 1
                        # Get available polish codes for suggestion
                        available_polish = list(PolishFinishType.objects.values_list('polish_internal', flat=True)[:5])
                        polish_suggestion = f" Available polish codes: {', '.join(available_polish)}" if available_polish else ""
                        failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not available in Master Data.{polish_suggestion}")
                        continue

                    # 5. Validate version code
                    version_obj = Version.objects.filter(version_internal=version_code).first()
                    if not version_obj:
                        failure_count += 1
                        # Get available version codes for suggestion
                        available_versions = list(Version.objects.values_list('version_internal', flat=True)[:5])
                        version_suggestion = f" Available version codes: {', '.join(available_versions)}" if available_versions else ""
                        failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not available in Master Data.{version_suggestion}")
                        continue

                    # 6. Validate category
                    category_obj = Category.objects.filter(category_name=category).first()
                    if not category_obj:
                        failure_count += 1
                        # Get available categories for suggestion
                        available_categories = list(Category.objects.values_list('category_name', flat=True)[:5])
                        category_suggestion = f" Available categories: {', '.join(available_categories)}" if available_categories else ""
                        failed_rows.append(f"Row {idx}: ❌ Category '{category}' not available in Master Data.{category_suggestion}")
                        continue
                    
                    print(f"✅ Category validation passed: '{category}' found in database")

                    # 7. Validate source format - handle both underscore and non-underscore formats
                    vendor_obj = None
                    location_obj = None

                    if "_" in source:
                        # Format: Vendor_Location (e.g., Titan_CPSE)
                        vendor_name, loc_name = source.split("_", 1)
                        
                        # Validate vendor by vendor_name field
                        vendor_obj = Vendor.objects.filter(vendor_name=vendor_name).first()
                        if not vendor_obj:
                            failure_count += 1
                            # Get available vendors for suggestion
                            available_vendors = list(Vendor.objects.values_list('vendor_name', flat=True)[:5])
                            vendor_suggestion = f" Available vendors: {', '.join(available_vendors)}" if available_vendors else ""
                            failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not available in Master Data.{vendor_suggestion}")
                            continue

                        # Validate location
                        location_obj = Location.objects.filter(location_name=loc_name).first()
                        if not location_obj:
                            failure_count += 1
                            # Get available locations for suggestion
                            available_locations = list(Location.objects.values_list('location_name', flat=True)[:5])
                            location_suggestion = f" Available locations: {', '.join(available_locations)}" if available_locations else ""
                            failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not available in Master Data.{location_suggestion}")
                            continue
                            
                        print(f"✅ Source with underscore - Vendor: '{vendor_name}' (Found: {vendor_obj.vendor_name}), Location: '{loc_name}'")
                    else:
                        # Format: Location only (e.g., CPSE, Mumbai, Delhi)
                        # Changed from Vendor to Location
                        location_obj = Location.objects.filter(location_name=source).first()
                        if not location_obj:
                            failure_count += 1
                            # Get available locations for suggestion
                            available_locations = list(Location.objects.values_list('location_name', flat=True)[:5])
                            location_suggestion = f" Available locations: {', '.join(available_locations)}" if available_locations else ""
                            failed_rows.append(f"Row {idx}: ❌ Location '{source}' not available in Master Data.{location_suggestion}")
                            continue
                            
                        # No vendor specified for this format
                        vendor_obj = None
                        print(f"✅ Source without underscore - Location only: '{source}' (Found: {location_obj.location_name})")

                    # Generate batch ID
                    batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{idx}"

                    # Create ModelMasterCreation record with validated category
                    ModelMasterCreation.objects.create(
                        batch_id=batch_id,
                        model_stock_no=model_stock,
                        plating_color=plating_obj_code.plating_color,  # Use resolved plating color
                        vendor_internal=vendor_obj.vendor_internal if vendor_obj else None,  # Handle None case
                        location=location_obj,  # Can be None for vendor-only format
                        tray_capacity=model_stock.tray_capacity if model_stock else None,
                        tray_type=model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                        ep_bath_type=model_stock.ep_bath_type if model_stock else None,
                        total_batch_quantity=input_qty,
                        version=version_obj if version_obj else None,
                        polish_finish=polish_obj if polish_obj else None,
                        category=category_obj,  # Save category object instead of string
                        plating_stk_no=plating_stock_no,           # <-- Save Plating Stk No
                        polishing_stk_no=polishing_stock_no,  # <-- Save Polishing Stk No   
                    )
                    print("✅ Row saved successfully!")
                    success_count += 1

                except Exception as e:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Error processing row: {str(e).splitlines()[0]}")
                    print(f"❌ Error processing row {idx}: {str(e)}")

            # Prepare response with detailed error list
            if success_count > 0 and failure_count == 0:
                return JsonResponse({
                    'success': True,
                    'message': f"✅ {success_count} row(s) processed successfully."
                })
            elif success_count > 0 and failure_count > 0:
                return JsonResponse({
                    'success': True,
                    'message': f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.",
                    'failed_rows': failed_rows  # This will show all the detailed error messages
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': f"❌ All {failure_count} row(s) failed to process.",
                    'failed_rows': failed_rows  # This will show all the detailed error messages
                }, status=400)

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': '❌ Invalid JSON data provided.'
            }, status=400)
        except Exception as e:
            print(f"❌ Error in datatable submission: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'❌ An error occurred: {str(e)}'
            }, status=500)

    def handle_file_upload(self, request):
        """Handle file upload with enhanced column validation"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "❌ No file uploaded.")
            return Response({'master_data': ModelMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid.")
            return Response({'master_data': ModelMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            if not sheet:
                messages.error(request, "❌ Could not read the Excel sheet.")
                return Response({'master_data': ModelMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

            # ========== NEW: VALIDATE COLUMNS BEFORE PROCESSING ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                messages.error(request, mark_safe(error_message.replace('\n', '<br>')))
                return Response({'master_data': ModelMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)
            
            print("✅ Excel column validation passed")
            # ========== END COLUMN VALIDATION ==========

            success_count = 0
            failure_count = 0
            failed_rows = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                # ========== ENHANCED: CHECK FOR MINIMUM REQUIRED COLUMNS ==========
                if len(row) < 7:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Missing Data — Expected 7 columns but found {len(row)}. Please ensure all required columns are filled.")
                    continue
                # ========== END ENHANCED COLUMN CHECK ==========

                # Extract data with new field structure
                s_no = str(row[0]).strip() if row[0] else ''
                plating_stock_no = str(row[1]).strip() if row[1] else ''
                polishing_stock_no = str(row[2]).strip() if row[2] else ''
                plating_colour = str(row[3]).strip() if row[3] else ''
                category = str(row[4]).strip() if row[4] else ''
                input_qty = row[5]
                source = str(row[6]).strip() if row[6] else ''

                # ========== ENHANCED: DETAILED EMPTY FIELD VALIDATION ==========
                empty_fields = []
                if not plating_stock_no:
                    empty_fields.append("Column B (Plating Stk No)")
                if not polishing_stock_no:
                    empty_fields.append("Column C (Polishing Stk No)")
                if not plating_colour:
                    empty_fields.append("Column D (Plating Colour)")
                if not category:
                    empty_fields.append("Column E (Category)")
                if input_qty in [None, '', 0]:
                    empty_fields.append("Column F (Input Qty)")
                if not source:
                    empty_fields.append("Column G (Source)")

                if empty_fields:
                    field_list = ", ".join(empty_fields)
                    failed_rows.append(f"Row {idx}: ❌ Empty Required Fields — {field_list} cannot be empty.")
                    failure_count += 1
                    continue
                # ========== END ENHANCED EMPTY FIELD VALIDATION ==========

                # Convert quantity to integer
                try:
                    input_qty = int(input_qty)
                except (ValueError, TypeError):
                    failed_rows.append(f"Row {idx}: ❌ Invalid Data — Column F (Input Qty): '{input_qty}' is not a valid number.")
                    failure_count += 1
                    continue

                # Validate codes using new stock numbers
                model_stock, codes, error_msg = self.validate_codes(plating_stock_no, polishing_stock_no)
                if error_msg:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: {error_msg}")
                    continue

                if not codes:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Could not extract codes from stock numbers.")
                    continue

                plating_color_internal, polish_code, version_code = codes
                print(f"🧪 Codes Extracted ➤ Plating Internal: '{plating_color_internal}', Polish: '{polish_code}', Version: '{version_code}'")

                # ========== APPLY SAME VALIDATION AS DATATABLE SUBMISSION ==========

                # 1. Validate plating code using the resolved plating_color_internal
                plating_obj_code = Plating_Color.objects.filter(plating_color_internal=plating_color_internal).first()
                if not plating_obj_code:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating color internal code '{plating_color_internal}' not available in Master Data.")
                    continue

                # 2. Validate plating color from input
                plating_color_obj = Plating_Color.objects.filter(plating_color=plating_colour).first()
                if not plating_color_obj:
                    failure_count += 1
                    # Get available plating colors for suggestion
                    available_colors = list(Plating_Color.objects.values_list('plating_color', flat=True)[:5])
                    color_suggestion = f" Available colors: {', '.join(available_colors)}" if available_colors else ""
                    failed_rows.append(f"Row {idx}: ❌ Plating Colour '{plating_colour}' not available in Master Data.{color_suggestion}")
                    continue

                # 3. Cross-validation: Check if resolved plating_color_internal matches with the plating color from Excel
                if plating_obj_code.pk != plating_color_obj.pk:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating color mismatch: Stock code '{plating_stock_no}' resolves to '{plating_obj_code.plating_color}' but Excel shows '{plating_colour}'.")
                    continue

                # 4. Validate polish code
                polish_obj = PolishFinishType.objects.filter(polish_internal=polish_code).first()
                if not polish_obj:
                    failure_count += 1
                    # Get available polish codes for suggestion
                    available_polish = list(PolishFinishType.objects.values_list('polish_internal', flat=True)[:5])
                    polish_suggestion = f" Available polish codes: {', '.join(available_polish)}" if available_polish else ""
                    failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not available in Master Data.{polish_suggestion}")
                    continue

                # 5. Validate version code
                version_obj = Version.objects.filter(version_internal=version_code).first()
                if not version_obj:
                    failure_count += 1
                    # Get available version codes for suggestion
                    available_versions = list(Version.objects.values_list('version_internal', flat=True)[:5])
                    version_suggestion = f" Available version codes: {', '.join(available_versions)}" if available_versions else ""
                    failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not available in Master Data.{version_suggestion}")
                    continue

                # 6. Validate category
                category_obj = Category.objects.filter(category_name=category).first()
                if not category_obj:
                    failure_count += 1
                    # Get available categories for suggestion
                    available_categories = list(Category.objects.values_list('category_name', flat=True)[:5])
                    category_suggestion = f" Available categories: {', '.join(available_categories)}" if available_categories else ""
                    failed_rows.append(f"Row {idx}: ❌ Category '{category}' not available in Master Data.{category_suggestion}")
                    continue

                # 7. Validate source format - handle both underscore and non-underscore formats
                vendor_obj = None
                location_obj = None

                if "_" in source:
                    # Format: Vendor_Location (e.g., Titan_CPSE)
                    vendor_name, loc_name = source.split("_", 1)
                    
                    # Validate vendor by vendor_name field
                    vendor_obj = Vendor.objects.filter(vendor_name=vendor_name).first()
                    if not vendor_obj:
                        failure_count += 1
                        # Get available vendors for suggestion
                        available_vendors = list(Vendor.objects.values_list('vendor_name', flat=True)[:5])
                        vendor_suggestion = f" Available vendors: {', '.join(available_vendors)}" if available_vendors else ""
                        failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not available in Master Data.{vendor_suggestion}")
                        continue

                    # Validate location
                    location_obj = Location.objects.filter(location_name=loc_name).first()
                    if not location_obj:
                        failure_count += 1
                        # Get available locations for suggestion
                        available_locations = list(Location.objects.values_list('location_name', flat=True)[:5])
                        location_suggestion = f" Available locations: {', '.join(available_locations)}" if available_locations else ""
                        failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not available in Master Data.{location_suggestion}")
                        continue
                        
                    print(f"✅ Source with underscore - Vendor: '{vendor_name}' (Found: {vendor_obj.vendor_name}), Location: '{loc_name}'")
                else:
                    # Format: Location only (e.g., CPSE, Mumbai, Delhi)
                    # Changed from Vendor to Location
                    location_obj = Location.objects.filter(location_name=source).first()
                    if not location_obj:
                        failure_count += 1
                        # Get available locations for suggestion
                        available_locations = list(Location.objects.values_list('location_name', flat=True)[:5])
                        location_suggestion = f" Available locations: {', '.join(available_locations)}" if available_locations else ""
                        failed_rows.append(f"Row {idx}: ❌ Location '{source}' not available in Master Data.{location_suggestion}")
                        continue
                        
                    # No vendor specified for this format
                    vendor_obj = None
                    print(f"✅ Source without underscore - Location only: '{source}' (Found: {location_obj.location_name})")

                # Generate batch ID
                batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{idx}"

                # Create ModelMasterCreation record with validated category
                ModelMasterCreation.objects.create(
                    batch_id=batch_id,
                    model_stock_no=model_stock,
                    plating_color=plating_obj_code.plating_color,  # Use resolved plating color
                    vendor_internal=vendor_obj.vendor_internal if vendor_obj else None,  # Handle None case
                    location=location_obj,  # Can be None for vendor-only format
                    tray_capacity=model_stock.tray_capacity if model_stock else None,
                    tray_type=model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                    ep_bath_type=model_stock.ep_bath_type if model_stock else None,
                    total_batch_quantity=input_qty,
                    version=version_obj if version_obj else None,
                    polish_finish=polish_obj if polish_obj else None,
                    category=category_obj,  # Save category object instead of string
                    plating_stk_no=plating_stock_no,           # <-- Save Plating Stk No
                    polishing_stk_no=polishing_stock_no,  # <-- Save Polishing Stk No   
                )
                print(f"✅ Row {idx} saved successfully!")
                success_count += 1

            # Return results with enhanced error messages
            if success_count > 0 and failure_count == 0:
                messages.success(request, f"✅ {success_count} row(s) uploaded successfully.")
            elif success_count > 0 and failure_count > 0:
                error_msg = f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.warning(request, mark_safe(error_msg.replace('\n', '<br>')))
            elif success_count == 0 and failure_count > 0:
                error_msg = f"❌ Upload Failed: All {failure_count} row(s) failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.error(request, mark_safe(error_msg.replace('\n', '<br>')))

            return Response({'master_data': ModelMasterCreation.objects.all()})

        except Exception as e:
            print(f"❌ Error processing file: {str(e)}")
            messages.error(request, f"❌ An error occurred: {str(e)}")
            return Response({'master_data': ModelMasterCreation.objects.all()}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class DPBulkUploadPreviewView(APIView):
    """
    Enhanced preview view with detailed column validation
    """
    parser_classes = [MultiPartParser]

    def validate_excel_columns(self, sheet):
        """
        Validate Excel file columns to ensure they match expected format
        Returns: (is_valid, error_message, header_row)
        """
        try:
            # Expected column names in exact order
            expected_columns = [
                'S.No',
                'Plating Stk No', 
                'Polishing Stk No',
                'Plating Colour',
                'Category',
                'Input Qty',
                'Source'
            ]
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_count = len(expected_columns) - len(actual_columns)
                missing_columns = expected_columns[len(actual_columns):]
                return False, f"❌ Missing {missing_count} column(s): {', '.join(missing_columns)}. Expected {len(expected_columns)} columns but found {len(actual_columns)}.", actual_columns
            
            # Check for exact column name matches
            mismatched_columns = []
            for i, (expected, actual) in enumerate(zip(expected_columns, actual_columns)):
                if expected != actual:
                    mismatched_columns.append(f"Column {i+1}: Expected '{expected}' but found '{actual}'")
            
            if mismatched_columns:
                error_msg = "❌ Column Name Mismatch:\n" + "\n".join(mismatched_columns)
                error_msg += f"\n\nExpected format: {' | '.join(expected_columns)}"
                error_msg += f"\nActual format: {' | '.join(actual_columns[:len(expected_columns)])}"
                return False, error_msg, actual_columns
            
            # Check for extra columns (warn but don't fail)
            if len(actual_columns) > len(expected_columns):
                extra_columns = actual_columns[len(expected_columns):]
                print(f"⚠️ Warning: Found {len(extra_columns)} extra column(s) that will be ignored: {', '.join(extra_columns)}")
            
            return True, "✅ Column validation passed", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'success': False, 'error': 'No file uploaded.'}, status=400)
        
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return Response({
                'success': False, 
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)
        
        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            if not sheet:
                return Response({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # ========== ENHANCED: USE DETAILED COLUMN VALIDATION ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                return Response({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            print("✅ Excel column validation passed for preview")
            # ========== END ENHANCED COLUMN VALIDATION ==========
            
            data = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not any(row):
                    continue
                
                # Check for minimum required columns
                if len(row) < 7:
                    continue  # Skip incomplete rows in preview
                
                data.append({
                    'S.No': row[0] or '',
                    'Plating Stk No': row[1] or '',
                    'Polishing Stk No': row[2] or '',
                    'Plating Colour': row[3] or '',
                    'Category': row[4] or '',
                    'Input Qty': row[5] or '',
                    'Source': row[6] or '',
                })
                
                # Limit preview to first 100 rows
                if len(data) >= 100:
                    break
            
            return Response({'success': True, 'data': data})
            
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=500)


# Alternative: Enhanced error messages with more specific column details
class DPBulkUploadPreviewViewEnhanced(APIView):
    """
    Even more enhanced preview view with ultra-specific column validation
    """
    parser_classes = [MultiPartParser]

    def validate_excel_columns_enhanced(self, sheet):
        """
        Enhanced validation with even more specific error messages
        """
        try:
            # Expected column names with their positions
            expected_columns = {
                1: 'S.No',
                2: 'Plating Stk No', 
                3: 'Polishing Stk No',
                4: 'Plating Colour',
                5: 'Category',
                6: 'Input Qty',
                7: 'Source'
            }
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_positions = list(range(len(actual_columns) + 1, len(expected_columns) + 1))
                missing_details = [f"Column {pos} should be '{expected_columns[pos]}'" for pos in missing_positions]
                return False, f"❌ Missing Required Columns:\n" + "\n".join(missing_details) + f"\n\nFound {len(actual_columns)} columns, expected {len(expected_columns)}.", actual_columns
            
            # Check each column position for exact matches
            specific_errors = []
            for pos in range(1, len(expected_columns) + 1):
                expected_name = expected_columns[pos]
                actual_name = actual_columns[pos - 1] if pos <= len(actual_columns) else ''
                
                if expected_name != actual_name:
                    specific_errors.append(
                        f"Column {pos} Name Mismatch: Expected '{expected_name}' but found '{actual_name}'"
                    )
            
            if specific_errors:
                error_msg = "❌ Column Name Mismatch Detected:\n\n" + "\n".join(specific_errors)
                error_msg += f"\n\n📋 Required Column Format:"
                for pos, name in expected_columns.items():
                    status = "✅" if pos <= len(actual_columns) and actual_columns[pos-1] == name else "❌"
                    current = actual_columns[pos-1] if pos <= len(actual_columns) else "MISSING"
                    error_msg += f"\n   Column {pos}: {name} {status} (Current: {current})"
                
                return False, error_msg, actual_columns
            
            return True, "✅ All column names validated successfully", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'success': False, 'error': 'No file uploaded.'}, status=400)
        
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return Response({
                'success': False, 
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)
        
        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            if not sheet:
                return Response({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # Use enhanced column validation
            is_valid, error_message, actual_columns = self.validate_excel_columns_enhanced(sheet)
            if not is_valid:
                return Response({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            # Rest of the processing logic remains the same...
            data = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not any(row):
                    continue
                
                if len(row) < 7:
                    continue
                
                data.append({
                    'S.No': row[0] or '',
                    'Plating Stk No': row[1] or '',
                    'Polishing Stk No': row[2] or '',
                    'Plating Colour': row[3] or '',
                    'Category': row[4] or '',
                    'Input Qty': row[5] or '',
                    'Source': row[6] or '',
                })
                
                if len(data) >= 100:
                    break
            
            return Response({'success': True, 'data': data})
            
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=500)
 
 
@method_decorator(csrf_exempt, name='dispatch')
class GetPlatingColourAPIView(APIView):
    """
    API endpoint to fetch plating colour based on color code from Plating Stk No
    
    POST /dayplanning/get_plating_colour/
    {
        "color_code": "S",
        "plating_stock_no": "1805SAB02"
    }
    
    Response:
    {
        "success": true,
        "plating_colour": "Silver",
        "color_code": "S"
    }
    """
    
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            color_code = data.get('color_code', '').strip().upper()
            plating_stock_no = data.get('plating_stock_no', '').strip()
            
            if not color_code:
                return JsonResponse({
                    'success': False,
                    'error': 'Color code is required'
                }, status=400)
            
            # Handle ambiguous color codes (with "/" identifier)
            plating_color_internal = color_code
            if "/" in plating_stock_no:
                # Extract the part after "/" for ambiguous resolution
                additional_identifier = plating_stock_no.split("/")[1]
                plating_color_internal = additional_identifier
                print(f"🔍 Using additional identifier for ambiguous color: {additional_identifier}")
            
            # Try to find plating color by internal code
            plating_color_obj = Plating_Color.objects.filter(
                plating_color_internal=plating_color_internal
            ).first()
            
            if plating_color_obj:
                return JsonResponse({
                    'success': True,
                    'plating_colour': plating_color_obj.plating_color,
                    'color_code': color_code,
                    'plating_color_internal': plating_color_internal
                })
            else:
                # Get available colors for suggestion
                available_colors = list(Plating_Color.objects.values_list(
                    'plating_color_internal', 'plating_color'
                )[:10])
                
                return JsonResponse({
                    'success': False,
                    'error': f"Plating color internal code '{plating_color_internal}' not found in Master Data.",
                    'available_colors': [
                        {'code': code, 'color': color} 
                        for code, color in available_colors
                    ]
                }, status=404)
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            print(f"❌ Error in GetPlatingColourAPIView: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetCategoriesAPIView(APIView):
    """
    API endpoint to fetch all categories for dropdown
    
    GET /dayplanning/get_categories/
    
    Response:
    {
        "success": true,
        "categories": [
            {"category_name": "Category A"},
            {"category_name": "Category B"}
        ]
    }
    """
    
    def get(self, request):
        try:
            # Import the Category model (you'll need to add this import at the top of views.py)
            from modelmasterapp.models import Category
            
            categories = list(Category.objects.values('category_name').order_by('category_name'))
            
            return JsonResponse({
                'success': True,
                'categories': categories,
                'count': len(categories)
            })
            
        except Exception as e:
            print(f"❌ Error in GetCategoriesAPIView: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetLocationsAPIView(APIView):
    """
    API endpoint to fetch all locations for dropdown
    
    GET /dayplanning/get_locations/
    
    Response:
    {
        "success": true,
        "locations": [
            {"location_name": "Location A"},
            {"location_name": "Location B"}
        ]
    }
    """
    
    def get(self, request):
        try:
            # Location model should already be imported
            locations = list(Location.objects.values('location_name').order_by('location_name'))
            
            return JsonResponse({
                'success': True,
                'locations': locations,
                'count': len(locations)
            })
            
        except Exception as e:
            print(f"❌ Error in GetLocationsAPIView: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            }, status=500) 


class DayPlanningPickTableAPIView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_PickTable.html'
    permission_classes = [IsAuthenticated] 

    def get(self, request, *args, **kwargs):
        user = request.user

        # Check if user is in Admin group
        is_admin = user.groups.filter(name='Admin').exists() if user.is_authenticated else False
        
        # Subqueries for annotations
        last_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]

        
        accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('accepted_Ip_stock')[:1]   

        # ✅ NEW: Add subquery to check if tray scanning was started but not completed
        tray_scan_status_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('tray_scan_status')[:1]

        # Build queryset
        queryset = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=False  
        ).annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),
            accepted_Ip_stock=Subquery(accepted_Ip_stock_subquery),
            tray_scan_status=Subquery(tray_scan_status_subquery),  # ✅ NEW
        ).order_by('-date_time')

        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)  # 10 items per page
        page_obj = paginator.get_page(page_number)

        # Convert page_obj to list of dicts
        master_data = list(page_obj.object_list.values(
            'batch_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'version__version_internal',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_D_Picker',
            'last_process_module',
            'next_process_module',
            'Draft_Saved',
            'dp_pick_remarks',
            'top_tray_qty_verified',
            'plating_stk_no',
            'polishing_stk_no',
            'category',
            'hold_lot',
            'release_lot',
            'holding_reason',
            'release_reason',
            'accepted_Ip_stock',
            'tray_scan_status',  # ✅ NEW
        ))

        # Calculate no_of_trays dynamically and determine needs_top_tray_scan
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            tray_capacity = data.get('tray_capacity', 0)
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"

            # ✅ NEW: Determine if this lot needs top tray scan
            # Condition: tray_scan_status=True but Moved_to_D_Picker=False
            tray_scan_status = data.get('tray_scan_status', False)
            moved_to_d_picker = data.get('Moved_to_D_Picker', False)
            
            # This indicates partial completion - user needs to set top tray
            data['needs_top_tray_scan'] = bool(tray_scan_status and not moved_to_d_picker)
            
            if tray_capacity > 0:
                no_of_trays = math.ceil(total_batch_quantity / tray_capacity)
                data['no_of_trays'] = no_of_trays
                tray_qty_list = []
                remainder = total_batch_quantity % tray_capacity
                if no_of_trays == 1:
                    tray_qty_list = [total_batch_quantity]
                elif no_of_trays > 1:
                    if remainder != 0:
                        tray_qty_list.append(remainder)
                        for _ in range(1, no_of_trays):
                            tray_qty_list.append(tray_capacity)
                    else:
                        for _ in range(no_of_trays):
                            tray_qty_list.append(tray_capacity)
                data['tray_qty_list'] = tray_qty_list
            else:
                data['no_of_trays'] = 0
                data['tray_qty_list'] = []
                
            # Add model images
            mmc = ModelMasterCreation.objects.filter(batch_id=data['batch_id']).first()
            images = []
            if mmc:
                model_master = mmc.model_stock_no
                for img in getattr(model_master, 'images', []).all():
                    if getattr(img, 'master_image', None):
                        images.append(img.master_image.url)
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.png')]
            data['model_images'] = images

        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
            'is_admin': is_admin,
        }
        return Response(context, template_name=self.template_name)



@method_decorator(csrf_exempt, name='dispatch')
class SaveHoldUnholdReasonAPIView(APIView):
    """
    POST with:
    {
        "batch_id": "BATCH-20240618123456-1",
        "remark": "Reason text",
        "action": "hold",  # or "unhold"
        "previous_status": "yet_to_start"  # optional, for restoring status
    }
    """
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            remark = data.get('remark', '').strip()
            action = data.get('action', '').strip().lower()
            previous_status = data.get('previous_status', '').strip()

            if not batch_id or not remark or action not in ['hold', 'unhold']:
                return JsonResponse({'success': False, 'error': 'Missing or invalid parameters.'}, status=400)

            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found.'}, status=404)

            if action == 'hold':
                # Store current lot status before holding
                if obj.Moved_to_D_Picker:
                    obj.previous_lot_status = 'yet_to_release'
                elif obj.Draft_Saved:
                    obj.previous_lot_status = 'draft'

                else:
                    obj.previous_lot_status = 'yet_to_start'
                
                obj.holding_reason = remark
                obj.hold_lot = True
                obj.release_reason = ''
                obj.release_lot = False
                
            elif action == 'unhold':
                obj.release_reason = remark
                obj.hold_lot = False
                obj.release_lot = True
                # previous_lot_status is preserved for frontend to restore status

            obj.save(update_fields=[
                'holding_reason', 'release_reason', 'hold_lot', 'release_lot', 'previous_lot_status'
            ])
            
            return JsonResponse({
                'success': True, 
                'message': 'Reason saved.',
                'previous_status': obj.previous_lot_status
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayIdScanAPIView(APIView):
    """
    Enhanced TrayId Scan API - ONLY ALLOW PRE-EXISTING TRAYS (no new tray creation)
    """

    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            user = request.user if request.user.is_authenticated else None

            lot_id = data.get('lot_id') or self.generate_new_lot_id()

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            # ✅ ENHANCED: Pre-validate all tray IDs - MUST EXIST IN SYSTEM
            tray_type_errors = []
            already_scanned_errors = []
            duplicate_tray_ids = []
            tray_not_in_system_errors = []  # ✅ NEW: Track trays not in system
            
            for i, tray in enumerate(trays):
                tray_id = tray.get('tray_id', '').strip()
                if not tray_id:
                    continue
                
                # ✅ NEW: First check if tray exists in TrayId table
                existing_tray = TrayId.objects.filter(tray_id=tray_id).first()
                
                if not existing_tray:
                    # ✅ NEW: Reject trays that don't exist in system
                    tray_not_in_system_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.'
                    })
                    continue
                
                # ✅ Check if tray is already scanned (and not delinked)
                if existing_tray.scanned and not existing_tray.delink_tray:
                    already_scanned_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'batch_info': existing_tray.batch_id.batch_id if existing_tray.batch_id else 'Unknown batch',
                        'scan_date': existing_tray.date.strftime('%d-%m-%Y %H:%M') if existing_tray.date else 'Unknown date'
                    })
                    continue
                
                # Check if tray belongs to different batch (and not delinked)
                if (existing_tray.batch_id and 
                    existing_tray.batch_id != batch_instance and 
                    not existing_tray.delink_tray):
                    duplicate_tray_ids.append(tray_id)
                    continue
                
                # Validate tray type for existing trays
                validation_helper = TrayIdUniqueCheckAPIView()
                validation_result = validation_helper.validate_tray_type_compatibility(
                    existing_tray, batch_id
                )
                
                if not validation_result['compatible']:
                    tray_type_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'error': validation_result['error'],
                        'batch_tray_type': validation_result['batch_tray_type'],
                        'scanned_tray_type': validation_result['scanned_tray_type']
                    })

            # ✅ NEW: Return tray not in system errors
            if tray_not_in_system_errors:
                error_messages = []
                for error in tray_not_in_system_errors:
                    error_messages.append(
                        f"Position {error['position']}: {error['tray_id']} - Not found in system"
                    )
                
                return JsonResponse({
                    'success': False,
                    'error': 'Some trays are not in the system',
                    'tray_not_in_system_errors': tray_not_in_system_errors,
                    'error_details': error_messages
                }, status=400)

            # Return already scanned errors if any
            if already_scanned_errors:
                error_messages = []
                for error in already_scanned_errors:
                    error_messages.append(
                        f"Position {error['position']}: {error['tray_id']} - Already scanned on {error['scan_date']}"
                    )
                
                return JsonResponse({
                    'success': False,
                    'error': 'Some trays are already scanned',
                    'already_scanned_errors': already_scanned_errors,
                    'error_details': error_messages
                }, status=400)

            # Return tray type errors if any
            if tray_type_errors:
                error_messages = []
                for error in tray_type_errors:
                    error_messages.append(f"Position {error['position']}: {error['error']}")
                
                return JsonResponse({
                    'success': False,
                    'error': 'Tray type validation failed',
                    'tray_type_errors': tray_type_errors,
                    'error_details': error_messages
                }, status=400)

            # Return duplicate errors if any
            if duplicate_tray_ids:
                return JsonResponse({
                    'success': False,
                    'error': f'Duplicate Tray ID(s) found: {", ".join(duplicate_tray_ids)}',
                    'duplicate_tray_ids': duplicate_tray_ids
                }, status=400)

            # Check if top tray (first tray) has quantity 0
            top_tray_qty_zero = False
            if trays and len(trays) > 0:
                first_tray_qty = int(trays[0].get('tray_quantity', 0))
                if first_tray_qty == 0:
                    top_tray_qty_zero = True
                    print(f"🔍 Top tray quantity is 0 - special handling required")

            # Fetch related fields from batch_instance
            model_stock_no = batch_instance.model_stock_no
            version = batch_instance.version
            total_batch_quantity = batch_instance.total_batch_quantity
            polish_finish_obj = PolishFinishType.objects.filter(polish_finish=batch_instance.polish_finish).first() 
            plating_color = batch_instance.plating_color

            # In the TrayIdScanAPIView post method, find this section and update it:

            with transaction.atomic():
                # Calculate total quantity excluding delinked trays (qty = 0)
                active_total_quantity = sum(int(tray.get('tray_quantity', 0)) for tray in trays if int(tray.get('tray_quantity', 0)) > 0)
                
                # Save TotalStockModel with active quantity only
                total_stock_obj = TotalStockModel.objects.create(
                    batch_id=batch_instance,
                    model_stock_no=model_stock_no,
                    version=version,
                    total_stock=active_total_quantity,
                    polish_finish=polish_finish_obj,
                    plating_color=Plating_Color.objects.filter(plating_color=plating_color).first() if plating_color else None,
                    lot_id=lot_id,
                    tray_scan_status=True,  # ✅ Mark as True to indicate tray scanning started
                    last_process_module="DayPlanning",
                    next_process_module="IP Screening",
                )

                # Process each tray - ONLY UPDATE EXISTING TRAYS
                for i, tray in enumerate(trays):
                    tray_id = tray.get('tray_id')
                    tray_quantity = tray.get('tray_quantity')
                    
                    if not tray_id or tray_quantity is None:
                        continue
                    
                    tray_quantity = int(tray_quantity)
                    
                    # Only set top_tray=True for first tray if its quantity > 0
                    if top_tray_qty_zero:
                        is_top_tray = False
                    else:
                        is_top_tray = (i == 0)
                    
                    # Check if tray should be delinked (quantity = 0)
                    is_delinked = (tray_quantity == 0)
                    delink_qty = None
                    
                    if is_delinked:
                        # Store original quantity from draft or previous value
                        draft_tray = DraftTrayId.objects.filter(
                            batch_id=batch_instance, 
                            tray_id=tray_id
                        ).first()
                        if draft_tray and draft_tray.delink_tray_qty:
                            delink_qty = draft_tray.delink_tray_qty
                        else:
                            delink_qty = str(batch_instance.tray_capacity or 0)
                    
                    # ✅ CHANGED: ONLY UPDATE EXISTING TRAYS (no new tray creation)
                    existing_tray = TrayId.objects.filter(tray_id=tray_id).first()
                    
                    if existing_tray:
                        # Update existing tray while preserving tray_type and tray_capacity
                        existing_tray.lot_id = lot_id
                        existing_tray.batch_id = batch_instance
                        existing_tray.tray_quantity = tray_quantity
                        existing_tray.user = user
                        existing_tray.delink_tray = is_delinked
                        existing_tray.delink_tray_qty = delink_qty
                        existing_tray.top_tray = is_top_tray
                        existing_tray.date = now()
                        existing_tray.scanned = True  # Set scanned status to True
                        existing_tray.new_tray = False  # <-- Set new_tray to False as requested

                        # Don't update tray_type and tray_capacity - preserve admin settings
                        if is_delinked:
                            existing_tray.delink_tray = True
                            existing_tray.lot_id = None
                            existing_tray.batch_id = None
                            existing_tray.scanned = False
                            existing_tray.IP_tray_verified = False
                            existing_tray.top_tray = False
                            existing_tray.save(update_fields=[
                                'delink_tray', 'lot_id', 'batch_id', 'scanned', 'IP_tray_verified', 'top_tray'
                            ])
                        else:
                            # Don't update tray_type and tray_capacity - preserve admin settings
                            existing_tray.delink_tray = False
                            existing_tray.save()
                        print(f"✅ Updated existing tray: {tray_id} (Top Tray: {is_top_tray}, Type: {existing_tray.tray_type}, Scanned: True)")
                    else:
                        # ✅ CHANGED: This should never happen now since we validate existence above
                        print(f"❌ ERROR: Tray {tray_id} not found in system - this should have been caught in validation")
                        return JsonResponse({
                            'success': False, 
                            'error': f'Tray ID {tray_id} not found in system'
                        }, status=400)
                

                # ✅ ENHANCED: Handle Moved_to_D_Picker based on top tray quantity
                if top_tray_qty_zero:
                    # ✅ NEW: If top tray quantity is zero, keep as draft mode (Moved_to_D_Picker=False)
                    # This allows user to complete the "Set Top Tray" step later
                    print(f"🔍 Top tray quantity is zero - keeping Moved_to_D_Picker=False (draft mode)")
                    # Do not update Moved_to_D_Picker, it remains False
                else:
                    # ✅ NORMAL: Complete tray scanning process
                    ModelMasterCreation.objects.filter(batch_id=batch_id).update(Moved_to_D_Picker=True)
                    print(f"✅ Normal tray scanning completed - set Moved_to_D_Picker=True")

            # Return response based on top tray quantity
            if top_tray_qty_zero:
                return JsonResponse({
                    'success': True, 
                    'message': 'Tray scan saved! Please scan top tray ID.',
                    'top_tray_scan_required': True,
                    'batch_id': batch_id,
                    'keep_modal_open': True
                }, status=201)
            else:
                return JsonResponse({'success': True, 'message': 'Tray scan and stock saved!'}, status=201)
        except IntegrityError as e:
            error_message = str(e)
            if 'duplicate key value violates unique constraint' in error_message and 'tray_id' in error_message:
                import re
                tray_match = re.search(r'Key \(tray_id\)=\(([^)]+)\)', error_message)
                duplicate_tray = tray_match.group(1) if tray_match else 'unknown'
                return JsonResponse({
                    'success': False, 
                    'error': f'Tray ID {duplicate_tray} already exists in the system.',
                    'duplicate_tray_ids': [duplicate_tray]
                }, status=400)
            else:
                return JsonResponse({'success': False, 'error': 'Database integrity error: ' + str(e)}, status=400)
        except Exception as e:
            print(f"❌ Error in TrayIdScanAPIView: {str(e)}")
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

    def generate_new_lot_id(self):
        from datetime import datetime
        timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
        last_lot = TotalStockModel.objects.order_by('-id').first()
        if last_lot and last_lot.lot_id and last_lot.lot_id.startswith("LID"):
            last_seq_no = int(last_lot.lot_id[-4:])
            next_seq_no = last_seq_no + 1
        else:
            next_seq_no = 1
        seq_no = f"{next_seq_no:04d}"
        return f"LID{timestamp}{seq_no}"

# Add this new API view to your views.py file

@method_decorator(csrf_exempt, name='dispatch')
class ValidateTopTrayAPIView(APIView):
    """
    Enhanced API endpoint for validating tray ID for top tray selection - ONLY ALLOW EXISTING TRAYS
    """
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            tray_id = data.get('tray_id', '').strip()

            if not batch_id or not tray_id:
                return JsonResponse({
                    'success': False,
                    'valid': False,
                    'error': 'Missing batch ID or tray ID'
                }, status=400)

            # Check if batch exists
            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({
                    'success': False,
                    'valid': False,
                    'error': 'Batch not found'
                }, status=404)

            # ✅ NEW: First check if tray exists in TrayId table at all
            tray_in_system = TrayId.objects.filter(tray_id=tray_id).first()
            if not tray_in_system:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.'
                })

            # Check if tray exists in this specific batch
            tray = TrayId.objects.filter(
                tray_id=tray_id,
                batch_id=batch_instance
            ).first()

            if not tray:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" not found in this batch'
                })

            # Check if the tray is delinked
            if tray.delink_tray:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" is delinked and cannot be set as top tray'
                })

            # Check if the tray has quantity 0
            if tray.tray_quantity <= 0:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" has zero quantity and cannot be set as top tray'
                })

            # Validate tray type compatibility
            validation_helper = TrayIdUniqueCheckAPIView()
            validation_result = validation_helper.validate_tray_type_compatibility(tray, batch_id)
            
            if not validation_result['compatible']:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': validation_result['error'],
                    'batch_tray_type': validation_result['batch_tray_type'],
                    'scanned_tray_type': validation_result['scanned_tray_type']
                })

            # All checks passed: Valid tray for top tray selection
            return JsonResponse({
                'success': True,
                'valid': True,
                'message': f'Tray ID "{tray_id}" is valid for top tray selection'
            })

        except Exception as e:
            print(f"❌ Error in ValidateTopTrayAPIView: {str(e)}")
            return JsonResponse({
                'success': False,
                'valid': False,
                'error': f'Validation error: {str(e)}'
            }, status=500)

# Update your TopTrayScanAPIView in views.py

@method_decorator(csrf_exempt, name='dispatch')
class TopTrayScanAPIView(APIView):
    """
    API endpoint to handle top tray scanning when original top tray quantity is 0
    Enhanced to complete the tray scanning process by setting Moved_to_D_Picker=True
    """
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            scanned_tray_id = data.get('scanned_tray_id', '').strip()
            user = request.user if request.user.is_authenticated else None

            if not batch_id or not scanned_tray_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'Missing batch ID or tray ID.'
                }, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({
                    'success': False, 
                    'error': 'Invalid batch ID.'
                }, status=404)

            with transaction.atomic():
                # Check if the scanned tray ID exists in the same batch
                existing_tray = TrayId.objects.filter(
                    tray_id=scanned_tray_id,
                    batch_id=batch_instance
                ).first()

                if not existing_tray:
                    return JsonResponse({
                        'success': False,
                        'error': f'Tray ID "{scanned_tray_id}" not found in this batch.'
                    }, status=400)

                # Check if this tray is delinked
                if existing_tray.delink_tray:
                    return JsonResponse({
                        'success': False,
                        'error': f'Tray ID "{scanned_tray_id}" is delinked and cannot be set as top tray.'
                    }, status=400)

                # ✅ UPDATED: Remove top_tray status from ALL trays in this batch first
                TrayId.objects.filter(
                    batch_id=batch_instance,
                    top_tray=True
                ).update(top_tray=False)
                
                print(f"🔄 Removed top_tray status from all trays in batch {batch_id}")

                # ✅ UPDATED: Set ONLY the scanned tray as the new top tray
                existing_tray.top_tray = True
                existing_tray.save(update_fields=['top_tray'])
                
                print(f"✅ Set {scanned_tray_id} as new top tray for batch {batch_id}")

                # ✅ NEW: Complete the tray scanning process by setting Moved_to_D_Picker=True
                # This indicates that the tray scanning is now fully completed
                batch_instance.Moved_to_D_Picker = True
                batch_instance.save(update_fields=['Moved_to_D_Picker'])
                
                print(f"✅ Set Moved_to_D_Picker=True for batch {batch_id} - tray scanning completed")

                return JsonResponse({
                    'success': True,
                    'message': f'Top tray set successfully: {scanned_tray_id}. Tray scanning completed!'
                })

        except Exception as e:
            print(f"❌ Error in TopTrayScanAPIView: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': f'An error occurred: {str(e)}'
            }, status=500)
            

@method_decorator(csrf_exempt, name='dispatch')
class VerifyTopTrayQtyAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            verified_tray_qty = data.get('verified_tray_qty')
            
            print(f"🔍 VerifyTopTrayQty called with batch_id={batch_id}, verified_tray_qty={verified_tray_qty}")
            
            if not batch_id or verified_tray_qty is None:
                return JsonResponse({'success': False, 'error': 'Missing batch_id or verified_tray_qty'}, status=400)
            
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)

            print(f"📋 Found batch: {obj.batch_id}")
            print(f"📊 Current values: total_batch_quantity={obj.total_batch_quantity}, initial_batch_quantity={obj.initial_batch_quantity}")

            # Store the original total_batch_quantity as initial_batch_quantity if not already set
            if not obj.initial_batch_quantity:
                obj.initial_batch_quantity = obj.total_batch_quantity
                print(f"💾 Set initial_batch_quantity = {obj.initial_batch_quantity}")

            # Update the verified quantity and verification status
            obj.verified_tray_qty = verified_tray_qty
            obj.total_batch_quantity = verified_tray_qty  # Update total to verified amount
            obj.top_tray_qty_verified = True
            obj.Draft_Saved = True
            
            # Save the changes
            obj.save(update_fields=[
                'top_tray_qty_verified', 
                'verified_tray_qty', 
                'total_batch_quantity', 
                'initial_batch_quantity',
                'Draft_Saved'
            ])
            
            print(f"✅ Successfully updated batch {batch_id}")
            print(f"📈 Final values: total_batch_quantity={obj.total_batch_quantity}, verified_tray_qty={obj.verified_tray_qty}, top_tray_qty_verified={obj.top_tray_qty_verified}")
            
            return JsonResponse({
                'success': True, 
                'message': 'Top tray quantity verified successfully.',
                'verified_tray_qty': obj.verified_tray_qty,
                'total_batch_quantity': obj.total_batch_quantity,
                'top_tray_qty_verified': obj.top_tray_qty_verified
            })
                     
        except Exception as e:
            print(f"❌ Error in VerifyTopTrayQtyAPIView: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ UPDATED: Get top tray first, then other trays
        # First, get the top tray (top_tray=True)
        top_tray = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False,
            top_tray=True
        ).first()
        
        # Then get all other trays (excluding the top tray)
        other_trays = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False,
            top_tray=False
        ).order_by('id')
        
        data = []
        
        # Add top tray first if it exists
        if top_tray:
            data.append({
                's_no': 1,  # Always 1 for top tray
                'tray_id': top_tray.tray_id,
                'tray_quantity': top_tray.tray_quantity,
                'position': 0,  # Top position
                'is_top_tray': True
            })
        
        # Add other trays starting from position 2
        for idx, tray in enumerate(other_trays):
            data.append({
                's_no': idx + 2,  # Start from 2 since top tray is 1
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx + 1,  # Position after top tray
                'is_top_tray': False
            })
        
        return JsonResponse({'success': True, 'trays': data})
    
    
    
@method_decorator(csrf_exempt, name='dispatch')
class DraftTrayIdAPIView(APIView):
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            user = request.user if request.user.is_authenticated else None

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            lot_id = data.get('lot_id') or f"DRAFT-{batch_id}"

            with transaction.atomic():
                # Always clear all existing draft entries for this batch before saving new ones
                DraftTrayId.objects.filter(batch_id=batch_instance).delete()

                for tray in trays:
                    tray_id = tray.get('tray_id', '').strip()
                    tray_quantity = tray.get('tray_quantity')
                    position = tray.get('position')
                    
                    if tray_quantity is None or position is None:
                        continue
                    
                    tray_quantity = int(tray_quantity)
                    is_delinked = (tray_quantity == 0)
                    delink_qty = None

                    DraftTrayId.objects.create(
                        lot_id=lot_id,
                        batch_id=batch_instance,
                        tray_id=tray_id,
                        tray_quantity=tray_quantity,
                        position=position,
                        user=user,
                        delink_tray=is_delinked,
                        delink_tray_qty=delink_qty
                    )
                    print(f"✅ Saved draft: Position {position}, Tray ID: '{tray_id}', Qty: {tray_quantity}, Delinked: {is_delinked}")

            batch_instance.Draft_Saved = True
            batch_instance.save(update_fields=['Draft_Saved'])

            return JsonResponse({'success': True, 'message': 'Draft saved!'}, status=201)

        except IntegrityError as e:
            print(f"❌ IntegrityError in DraftTrayIdAPIView: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': f'Database constraint error: {str(e)}. Please refresh and try again.'
            }, status=400)
        except Exception as e:
            print(f"❌ Error in DraftTrayIdAPIView: {str(e)}")
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class DraftTrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ UPDATED: Filter out delinked trays and order by position
        draft_trays = DraftTrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False  # Only get non-delinked trays
        ).order_by('position')
        
        data = [
            {
                'id': tray.id,
                's_no': idx + 1,  # Reindex for display
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx  # Reset positions for non-delinked trays
            }
            for idx, tray in enumerate(draft_trays)
        ]
        
        return JsonResponse({'success': True, 'trays': data})
    
@method_decorator(csrf_exempt, name='dispatch')
class TrayIdUniqueCheckAPIView(APIView):
    """
    Enhanced TrayId validation - ONLY ALLOW PRE-EXISTING TRAYS from TrayId table
    """
    def get(self, request):
        tray_id = request.GET.get('tray_id')
        batch_id = request.GET.get('batch_id')
        
        if not tray_id:
            return JsonResponse({'exists': False, 'error': 'Missing tray_id'})
        
        # ✅ NEW: Check if tray exists in TrayId table FIRST
        existing_tray = TrayId.objects.filter(tray_id=tray_id).first()
        
        if not existing_tray:
            # ✅ CHANGED: Reject tray IDs that don't exist in the system
            return JsonResponse({
                'exists': False,
                'available': False,
                'tray_not_in_system': True,
                'delink_tray': None,  # <-- Add this line
                'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.',
                'message': 'This tray must be added by admin before scanning.'
            })
        
        # ✅ NEW: Disallow if tray is rejected
        if getattr(existing_tray, 'rejected_tray', False):
            return JsonResponse({
                'exists': True,
                'available': False,
                'rejected_tray': True,
                'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
                'error': f'Tray ID \"{tray_id}\" is marked as rejected and cannot be used.',
                'message': 'This tray is rejected and cannot be used for scanning.'
            })
        

        # Check if tray is delinked (can be reused regardless of scanned status)
        if existing_tray.delink_tray:
            # Delinked trays can be reused - validate tray type if batch_id provided
            if batch_id:
                validation_result = self.validate_tray_type_compatibility(existing_tray, batch_id)
                if not validation_result['compatible']:
                    return JsonResponse({
                        'exists': True,
                        'available': False,
                        'tray_type_error': True,
                        'delink_tray': True,  # <-- Add this line
                        'error': validation_result['error'],
                        'batch_tray_type': validation_result['batch_tray_type'],
                        'scanned_tray_type': validation_result['scanned_tray_type']
                    })
            
            return JsonResponse({
                'exists': True,
                'available': True,
                'delink_tray': True,  # <-- Add this line
                'status': 'delinked_reusable',
                'message': 'Delinked tray - available for reuse'
            })
        
        # Check if tray is already scanned/used
        if existing_tray.scanned:
            return JsonResponse({
                'exists': True,
                'available': False,
                'already_scanned': True,
                'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
                'error': f'Tray ID "{tray_id}" has already been scanned and is in use',
                'batch_info': existing_tray.batch_id.batch_id if existing_tray.batch_id else 'Unknown batch',
                'scan_date': existing_tray.date.strftime('%d-%m-%Y %H:%M') if existing_tray.date else 'Unknown date'
            })
        
        # Tray exists but not scanned - validate tray type compatibility
        if batch_id:
            validation_result = self.validate_tray_type_compatibility(existing_tray, batch_id)
            if not validation_result['compatible']:
                return JsonResponse({
                    'exists': True,
                    'available': False,
                    'tray_type_error': True,
                    'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
                    'error': validation_result['error'],
                    'batch_tray_type': validation_result['batch_tray_type'],
                    'scanned_tray_type': validation_result['scanned_tray_type']
                })
        
        # Tray is available for use
        return JsonResponse({
            'exists': True,
            'available': True,
            'status': 'pre_configured',
            'message': 'Pre-configured tray - available for scanning',
            'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
            'tray_type': existing_tray.tray_type,
            'tray_capacity': existing_tray.tray_capacity
        })
    
    def validate_tray_type_compatibility(self, tray, batch_id):
        """
        Validate if the scanned tray type is compatible with the batch tray type
        
        Args:
            tray: TrayId object with tray_type and tray_capacity
            batch_id: Batch ID string
            
        Returns:
            dict: {
                'compatible': bool,
                'error': str,
                'batch_tray_type': str,
                'scanned_tray_type': str
            }
        """
        try:
            # Get batch instance
            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return {
                    'compatible': False,
                    'error': 'Batch not found',
                    'batch_tray_type': None,
                    'scanned_tray_type': tray.tray_type
                }
            
            # Get batch tray type
            batch_tray_type = batch_instance.tray_type
            scanned_tray_type = tray.tray_type
            
            print(f"🔍 Tray Type Validation: Batch={batch_tray_type}, Scanned Tray={scanned_tray_type}")
            
            # If either tray type is not set, allow but warn
            if not batch_tray_type or not scanned_tray_type:
                return {
                    'compatible': True,  # Allow if not configured
                    'error': None,
                    'batch_tray_type': batch_tray_type,
                    'scanned_tray_type': scanned_tray_type
                }
            
            # Compare tray types (case-insensitive)
            if batch_tray_type.lower() != scanned_tray_type.lower():
                error_msg = f"❌ Tray Type Mismatch: Current batch requires '{batch_tray_type}' tray type, but scanned tray '{tray.tray_id}' is of type '{scanned_tray_type}'"
                return {
                    'compatible': False,
                    'error': error_msg,
                    'batch_tray_type': batch_tray_type,
                    'scanned_tray_type': scanned_tray_type
                }
            
            # Validate tray capacity compatibility (optional but recommended)
            batch_tray_capacity = batch_instance.tray_capacity
            scanned_tray_capacity = tray.tray_capacity
            
            if batch_tray_capacity and scanned_tray_capacity:
                if abs(batch_tray_capacity - scanned_tray_capacity) > 0:
                    error_msg = f"⚠️ Tray Capacity Mismatch: Batch capacity is {batch_tray_capacity}, but scanned tray capacity is {scanned_tray_capacity}"
                    return {
                        'compatible': False,
                        'error': error_msg,
                        'batch_tray_type': batch_tray_type,
                        'scanned_tray_type': scanned_tray_type
                    }
            
            # All validations passed
            return {
                'compatible': True,
                'error': None,
                'batch_tray_type': batch_tray_type,
                'scanned_tray_type': scanned_tray_type
            }
            
        except Exception as e:
            print(f"❌ Error in tray type validation: {str(e)}")
            return {
                'compatible': False,
                'error': f'Validation error: {str(e)}',
                'batch_tray_type': None,
                'scanned_tray_type': tray.tray_type if tray else None
            }



# Add this new API view to your views.py file

@method_decorator(csrf_exempt, name='dispatch')
class UpdateBatchQuantityAndColorAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            new_quantity = data.get('total_batch_quantity')
            new_plating_color = data.get('plating_color')
            
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id, Moved_to_D_Picker=False).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found or already moved'}, status=404)
            
            # Update quantity if provided
            if new_quantity is not None:
                try:
                    new_quantity = int(new_quantity)
                    if new_quantity <= 0:
                        return JsonResponse({'success': False, 'error': 'Quantity must be greater than 0'}, status=400)
                    obj.total_batch_quantity = new_quantity
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'error': 'Invalid quantity value'}, status=400)
            
            # Update plating color if provided
            if new_plating_color:
                # Validate that the plating color exists in Plating_Color table
                plating_color_obj = Plating_Color.objects.filter(plating_color=new_plating_color).first()
                if not plating_color_obj:
                    available_colors = list(Plating_Color.objects.values_list('plating_color', flat=True)[:5])
                    return JsonResponse({
                        'success': False, 
                        'error': f'Plating color "{new_plating_color}" not found in Master Data',
                        'available_colors': available_colors
                    }, status=400)
                obj.plating_color = new_plating_color
            
            # Save the changes
            update_fields = []
            if new_quantity is not None:
                update_fields.append('total_batch_quantity')
            if new_plating_color:
                update_fields.append('plating_color')
            
            if update_fields:
                obj.save(update_fields=update_fields)
            
            return JsonResponse({
                'success': True, 
                'message': 'Record updated successfully',
                'updated_quantity': obj.total_batch_quantity,
                'updated_plating_color': obj.plating_color
            })
            
        except Exception as e:
            print(f"❌ Error in UpdateBatchQuantityAndColorAPIView: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetPlatingColorsAPIView(APIView):
    """
    API endpoint to fetch all plating colors for dropdown navigation
    """
    def get(self, request):
        try:
            plating_colors = list(Plating_Color.objects.values_list('plating_color', flat=True).order_by('plating_color'))
            return JsonResponse({
                'success': True,
                'plating_colors': plating_colors,
                'count': len(plating_colors)
            })
        except Exception as e:
            print(f"❌ Error in GetPlatingColorsAPIView: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class DeleteBatchAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.delete()
            return JsonResponse({'success': True, 'message': 'Batch deleted'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class SaveDPPickRemarkAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            remark = data.get('remark', '').strip()
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.dp_pick_remarks = remark
            obj.save(update_fields=['dp_pick_remarks'])
            return JsonResponse({'success': True, 'message': 'Remark saved'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayValidateAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id_input = str(data.get('batch_id')).strip()
            tray_id = str(data.get('tray_id')).strip()
            print(f"[TrayValidateAPIView] User entered: batch_id={batch_id_input}, tray_id={tray_id}")

            # Try to match batch_id by code part (after last '-')
            trays = TrayId.objects.filter(batch_id__batch_id__icontains=batch_id_input)
            print(f"A   ll tray_ids for batch containing '{batch_id_input}': {[t.tray_id for t in trays]}")

            exists = trays.filter(tray_id=tray_id).exists()
            print(f"Exists in TrayId table? {exists}")

            return JsonResponse({'success': True, 'exists': exists})
        except Exception as e:
            print(f"[TrayValidateAPIView] Error: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
     
     
import pytz

class DPCompletedTableView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_Completed_Table.html'
    permission_classes = [IsAuthenticated] 

    def get(self, request, *args, **kwargs):
        from django.utils import timezone
        from datetime import datetime, timedelta

        user = request.user
        tz = pytz.timezone("Asia/Kolkata")
        now_local = timezone.now().astimezone(tz)
        today = now_local.date()
        yesterday = today - timedelta(days=1)

        # --- Use created_at for date filtering ---
        # Get all related created_at values for completed batches
        completed_batches = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=True
        ).values_list('batch_id', flat=True)

        # Get min/max created_at from TotalStockModel for these batches
        created_at_qs = TotalStockModel.objects.filter(
            batch_id__batch_id__in=completed_batches
        )
        min_created_at = created_at_qs.order_by('created_at').values_list('created_at', flat=True).first()
        max_created_at = created_at_qs.order_by('-created_at').values_list('created_at', flat=True).first()

        # Always use current date in IST
        today = now_local.date()
        yesterday = today - timedelta(days=1) 

        # Get date filter parameters from request
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        # Calculate date range
        if from_date_str and to_date_str:
            try:
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            except ValueError:
                from_date = yesterday
                to_date = today
        else:
            from_date = yesterday
            to_date = today

        # Convert dates to datetime objects for filtering (include full day)
        from_datetime = timezone.make_aware(datetime.combine(from_date, datetime.min.time()))
        to_datetime = timezone.make_aware(datetime.combine(to_date, datetime.max.time()))

        # Subqueries for annotations
        last_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]
        created_at_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('created_at')[:1]
        accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('accepted_Ip_stock')[:1]
        few_cases_accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('few_cases_accepted_Ip_stock')[:1]
        rejected_ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('rejected_ip_stock')[:1]

        # --- Filter by created_at in TotalStockModel ---
        # Get batch_ids where created_at is in range
        batch_ids_in_range = list(
            TotalStockModel.objects.filter(
                created_at__range=(from_datetime, to_datetime)
            ).values_list('batch_id__batch_id', flat=True)
        )

        queryset = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=True,
            batch_id__in=batch_ids_in_range
        ).annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),
            created_at=Subquery(created_at_subquery),
            accepted_Ip_stock=Subquery(accepted_Ip_stock_subquery),
            few_cases_accepted_Ip_stock=Subquery(few_cases_accepted_Ip_stock_subquery),
            rejected_ip_stock=Subquery(rejected_ip_stock_subquery),
        ).order_by('-created_at')

        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)
        page_obj = paginator.get_page(page_number)

        master_data = list(page_obj.object_list.values(
            'batch_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_D_Picker',
            'last_process_module',
            'next_process_module',
            'Draft_Saved',
            'top_tray_qty_verified',
            'created_at',
            'plating_stk_no',
            'polishing_stk_no',
            'category',
            'version__version_internal',
            'dp_pick_remarks',
            'accepted_Ip_stock',
            'rejected_ip_stock',
            'few_cases_accepted_Ip_stock',
        ))

        # Calculate no_of_trays dynamically and add tray_qty_list
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            tray_capacity = data.get('tray_capacity', 0)
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"
            
            if tray_capacity > 0:
                no_of_trays = math.ceil(total_batch_quantity / tray_capacity)
                data['no_of_trays'] = no_of_trays
                
                # Calculate tray_qty_list (same logic as DP_PickTable)
                tray_qty_list = []
                remainder = total_batch_quantity % tray_capacity
                if no_of_trays == 1:
                    tray_qty_list = [total_batch_quantity]
                elif no_of_trays > 1:
                    if remainder != 0:
                        tray_qty_list.append(remainder)
                        for _ in range(1, no_of_trays):
                            tray_qty_list.append(tray_capacity)
                    else:
                        for _ in range(no_of_trays):
                            tray_qty_list.append(tray_capacity)
                data['tray_qty_list'] = tray_qty_list
            else:
                data['no_of_trays'] = 0
                data['tray_qty_list'] = []
                
            # Add model images
            mmc = ModelMasterCreation.objects.filter(batch_id=data['batch_id']).first()
            images = []
            if mmc:
                model_master = mmc.model_stock_no
                for img in getattr(model_master, 'images', []).all():
                    if getattr(img, 'master_image', None):
                        images.append(img.master_image.url)
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.png')]
            data['model_images'] = images

        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
            'from_date': from_date.strftime('%Y-%m-%d'),  # 🔥 NEW: Pass dates to template
            'to_date': to_date.strftime('%Y-%m-%d'),      # 🔥 NEW: Pass dates to template
            'date_filter_applied': bool(from_date_str and to_date_str),  # 🔥 NEW: Flag to show if custom dates used
        }
        return Response(context, template_name=self.template_name)
    
    
    
@method_decorator(csrf_exempt, name='dispatch')
class CompletedTrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ UPDATED: Get top tray first, then other trays, and filter out trays with quantity 0
        top_tray = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            top_tray=True,
            tray_quantity__gt=0  # Only include if quantity > 0
        ).first()
        
        other_trays = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            top_tray=False,
            tray_quantity__gt=0  # Only include if quantity > 0
        ).order_by('id')
        
        data = []
        
        # Add top tray first if it exists
        if top_tray:
            data.append({
                's_no': 1,  # Always 1 for top tray
                'tray_id': top_tray.tray_id,
                'tray_quantity': top_tray.tray_quantity,
                'position': 0,  # Top position
                'is_top_tray': True
            })
        
        # Add other trays starting from position 2
        for idx, tray in enumerate(other_trays):
            data.append({
                's_no': idx + 2,  # Start from 2 since top tray is 1
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx + 1,  # Position after top tray
                'is_top_tray': False
            })
        
        return JsonResponse({'success': True, 'trays': data})
   
   
   
@method_decorator(csrf_exempt, name='dispatch')
class TrayAutoSaveAPIView(APIView):
    """
    API for cross-browser tray auto-save functionality
    """
    
    def post(self, request):
        """Save auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            tray_data = data.get('tray_data', [])
            modal_data = data.get('modal_data', {})
            
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Check if there's valid data to save
            has_valid_data = any(
                tray.get('trayId', '').strip() or tray.get('trayQty', '')
                for tray in tray_data
            )
            
            if not has_valid_data:
                # No valid data - delete any existing auto-save
                TrayAutoSaveData.objects.filter(user=request.user, batch_id=batch_id).delete()
                return JsonResponse({'success': True, 'message': 'Auto-save cleared (no data)'})
            
            # Prepare auto-save data
            auto_save_data = {
                'timestamp': timezone.now().isoformat(),
                'batch_id': batch_id,
                'tray_data': tray_data,
                'modal_data': modal_data,
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:200],  # Track browser
            }
            
            # Update or create auto-save record
            auto_save_obj, created = TrayAutoSaveData.objects.update_or_create(
                user=request.user,
                batch_id=batch_id,
                defaults={'auto_save_data': auto_save_data}
            )
            
            action = 'created' if created else 'updated'
            print(f"✅ Auto-save {action} for user {request.user.username}, batch {batch_id}")
            
            return JsonResponse({
                'success': True, 
                'message': f'Auto-save {action}',
                'data_id': auto_save_obj.id
            })
            
        except Exception as e:
            print(f"❌ Error in TrayAutoSaveAPIView POST: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    def get(self, request):
        """Retrieve auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            batch_id = request.GET.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Get auto-save data for this user and batch
            auto_save_obj = TrayAutoSaveData.objects.filter(
                user=request.user, 
                batch_id=batch_id
            ).first()
            
            if not auto_save_obj:
                return JsonResponse({
                    'success': True, 
                    'has_data': False, 
                    'message': 'No auto-save data found'
                })
            
            # Check if data is expired (older than 24 hours)
            if auto_save_obj.is_expired(hours=24):
                auto_save_obj.delete()
                return JsonResponse({
                    'success': True, 
                    'has_data': False, 
                    'message': 'Auto-save data expired and cleared'
                })
            
            # Return the auto-save data
            return JsonResponse({
                'success': True,
                'has_data': True,
                'data': auto_save_obj.auto_save_data,
                'saved_at': auto_save_obj.updated_at.isoformat(),
                'user': request.user.username
            })
            
        except Exception as e:
            print(f"❌ Error in TrayAutoSaveAPIView GET: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    def delete(self, request):
        """Clear auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            batch_id = request.GET.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Delete auto-save data
            deleted_count, _ = TrayAutoSaveData.objects.filter(
                user=request.user, 
                batch_id=batch_id
            ).delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Auto-save cleared ({deleted_count} records deleted)'
            })
            
        except Exception as e:
            print(f"❌ Error in TrayAutoSaveAPIView DELETE: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class TrayAutoSaveCleanupAPIView(APIView):
    """
    API to clean up old auto-save data (optional - can be called periodically)
    """
    
    def post(self, request):
        """Clean up auto-save data older than specified hours"""
        try:
            if not request.user.is_authenticated or not request.user.is_staff:
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
            
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            hours = data.get('hours', 72)  # Default: 72 hours (3 days)
            
            # Delete old auto-save data
            from datetime import timedelta
            cutoff_time = timezone.now() - timedelta(hours=hours)
            
            deleted_count, _ = TrayAutoSaveData.objects.filter(
                updated_at__lt=cutoff_time
            ).delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Cleaned up {deleted_count} old auto-save records (older than {hours} hours)'
            })
            
        except Exception as e:
            print(f"❌ Error in TrayAutoSaveCleanupAPIView: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
